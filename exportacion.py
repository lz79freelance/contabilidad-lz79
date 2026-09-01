"""
exportacion.py
--------------
Módulo completo de exportación contable para gestoría:
- Exportación e importación de Inventario en Excel (.xlsx)
- Libro de Facturas Emitidas y Recibidas en Excel (.xlsx)
- Resumen Fiscal Trimestral (Modelos 303 de IVA y 130 de IRPF) en HTML y Excel
"""

import os
import html
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from almacenamiento import get_ruta_plantillas, abrir_archivo_si_escritorio
from db import (
    obtener_inventario_db,
    guardar_producto_inventario_db,
    obtener_resumen_trimestre_db,
)


def estilo_encabezado_excel(ws, fila_num, columnas_total):
    """Estilos profesionales para cabeceras en Excel."""
    fuente_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    borde_fino = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    return fuente_header, fill_header, borde_fino


def exportar_inventario_excel() -> str:
    """Exporta el inventario completo a Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"
    ws.views.sheetView[0].showGridLines = True

    headers = ["SKU / Código", "Descripción del Producto", "Categoría", "Stock Actual", "Nivel Mínimo", "Costo Unitario (€)", "Valor Total (€)", "Estado"]
    ws.append(headers)

    fuente_h, fill_h, borde_f = estilo_encabezado_excel(ws, 1, len(headers))
    for col_idx in range(1, len(headers) + 1):
        celda = ws.cell(row=1, column=col_idx)
        celda.font = fuente_h
        celda.fill = fill_h
        celda.alignment = Alignment(horizontal="center", vertical="center")
        celda.border = borde_f

    items = obtener_inventario_db()
    for row_idx, (sku, desc, cat, stock, minimo, costo) in enumerate(items, start=2):
        val_total = stock * costo
        estado = "OK" if stock >= minimo else "REPOSICIÓN"
        ws.append([sku, desc, cat, stock, minimo, costo, val_total, estado])

        ws.cell(row=row_idx, column=4).alignment = Alignment(horizontal="right")
        ws.cell(row=row_idx, column=5).alignment = Alignment(horizontal="right")
        ws.cell(row=row_idx, column=6).number_format = "#,##0.00 €"
        ws.cell(row=row_idx, column=7).number_format = "#,##0.00 €"
        
        celda_est = ws.cell(row=row_idx, column=8)
        celda_est.alignment = Alignment(horizontal="center")
        if estado == "REPOSICIÓN":
            celda_est.font = Font(color="DC2626", bold=True)
            celda_est.fill = PatternFill(start_color="FEE2E2", fill_type="solid")
        else:
            celda_est.font = Font(color="16A34A", bold=True)

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 4, 12)

    ruta = os.path.join(get_ruta_plantillas(), "Inventario_LZ79.xlsx")
    wb.save(ruta)
    abrir_archivo_si_escritorio(os.path.abspath(ruta))
    return ruta


def importar_inventario_excel() -> tuple[bool, str]:
    """Importa o actualiza productos desde Inventario_LZ79.xlsx."""
    ruta = os.path.join(get_ruta_plantillas(), "Inventario_LZ79.xlsx")
    if not os.path.exists(ruta):
        return False, "No se encontró el archivo Inventario_LZ79.xlsx en la carpeta de plantillas."

    try:
        wb = load_workbook(ruta, data_only=True)
        ws = wb.active
        actualizados = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            sku = str(row[0]).strip()
            desc = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            cat = str(row[2]).strip() if len(row) > 2 and row[2] else "General"
            stock = int(row[3]) if len(row) > 3 and row[3] is not None else 0
            minimo = int(row[4]) if len(row) > 4 and row[4] is not None else 0
            costo = float(row[5]) if len(row) > 5 and row[5] is not None else 0.0

            if sku and desc:
                if guardar_producto_inventario_db(sku, desc, cat, stock, minimo, costo):
                    actualizados += 1

        return True, f"✅ Sincronizados {actualizados} artículos con éxito."
    except Exception as e:
        return False, f"⚠️ Error al importar Excel: {str(e)}"


def exportar_informe_gestoria_excel(anio: int, trimestre: int) -> str:
    """Exporta el libro oficial de IVA y balance trimestral para la gestoría en formato Excel multi-pestaña."""
    movimientos, totales = obtener_resumen_trimestre_db(anio, trimestre)

    wb = Workbook()

    # 1. Pestaña de Resumen Fiscal
    ws_resumen = wb.active
    ws_resumen.title = f"Resumen Fiscal {trimestre}T"
    ws_resumen.views.sheetView[0].showGridLines = True

    ws_resumen.append(["CIERRE FISCAL PARA GESTORÍA / ASESORÍA"])
    ws_resumen.append([f"Periodo: {trimestre}º Trimestre del {anio}", "", f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"])
    ws_resumen.append([])

    ws_resumen.append(["MODELO 303 - IMPUESTO SOBRE EL VALOR AÑADIDO (IVA)", ""])
    ws_resumen.append(["Base Imponible de Ventas / Ingresos (+):", totales["base_ventas"]])
    ws_resumen.append(["Cuota IVA Repercutido (+):", totales["iva_repercutido"]])
    ws_resumen.append(["Base Imponible de Compras / Gastos (-):", totales["base_gastos"]])
    ws_resumen.append(["Cuota IVA Soportado Deducible (-):", totales["iva_soportado"]])
    ws_resumen.append(["RESULTADO ESTIMADO MODELO 303:", totales["modelo_303"]])
    ws_resumen.append([])

    ws_resumen.append(["MODELO 130 - PAGO FRACCIONADO IRPF", ""])
    ws_resumen.append(["Rendimiento Neto Trimestre (Ingresos - Gastos):", totales["rendimiento_neto"]])
    ws_resumen.append(["Retenciones IRPF Soportadas en Facturas Emitidas (-):", totales["retenciones_ventas"]])
    ws_resumen.append(["PAGO FRACCIONADO ESTIMADO (20%):", totales["modelo_130"]])

    ws_resumen.column_dimensions["A"].width = 55
    ws_resumen.column_dimensions["B"].width = 22
    for fila in range(4, 16):
        celda_b = ws_resumen.cell(row=fila, column=2)
        if isinstance(celda_b.value, (int, float)):
            celda_b.number_format = "#,##0.00 €"
            celda_b.font = Font(name="Segoe UI", bold=True)

    # 2. Pestaña de Facturas Emitidas
    ws_ventas = wb.create_sheet(title="Facturas Emitidas")
    ws_ventas.views.sheetView[0].showGridLines = True
    headers_v = ["Nº Factura", "Fecha", "Cliente / Razón Social", "Concepto", "Base Imponible", "% IVA", "Cuota IVA", "% IRPF", "Retención IRPF", "Total Factura", "Estado"]
    ws_ventas.append(headers_v)

    fuente_h, fill_h, borde_f = estilo_encabezado_excel(ws_ventas, 1, len(headers_v))
    for c in range(1, len(headers_v) + 1):
        cel = ws_ventas.cell(row=1, column=c)
        cel.font = fuente_h
        cel.fill = fill_h
        cel.alignment = Alignment(horizontal="center")

    # 3. Pestaña de Facturas Recibidas / Gastos
    ws_gastos = wb.create_sheet(title="Facturas Recibidas")
    ws_gastos.views.sheetView[0].showGridLines = True
    headers_g = ["Nº Registro", "Fecha", "Proveedor / Acreedor", "Concepto", "Base Imponible", "% IVA", "Cuota IVA", "% IRPF", "Retención IRPF", "Total Gasto", "Estado"]
    ws_gastos.append(headers_g)
    for c in range(1, len(headers_g) + 1):
        cel = ws_gastos.cell(row=1, column=c)
        cel.font = fuente_h
        cel.fill = PatternFill(start_color="7F1D1D", end_color="7F1D1D", fill_type="solid")
        cel.alignment = Alignment(horizontal="center")

    fila_v, fila_g = 2, 2
    for m in movimientos:
        num_doc, tipo, fecha, cli, conc = m[1], m[2], m[3][:10], m[4] or "-", m[5] or "-"
        base, tipo_iva, c_iva = m[6] or 0.0, m[7] or 21.0, m[8] or 0.0
        tipo_irpf, c_irpf, total, estado = m[9] or 0.0, m[10] or 0.0, m[12] or 0.0, m[13] or "COBRADO"

        fila_datos = [num_doc, fecha, cli, conc, base, f"{tipo_iva:.0f}%", c_iva, f"{tipo_irpf:.0f}%", c_irpf, total, estado]

        if "Facturas" in tipo or "Ventas" in tipo:
            ws_ventas.append(fila_datos)
            ws_ventas.cell(row=fila_v, column=5).number_format = "#,##0.00 €"
            ws_ventas.cell(row=fila_v, column=7).number_format = "#,##0.00 €"
            ws_ventas.cell(row=fila_v, column=9).number_format = "#,##0.00 €"
            ws_ventas.cell(row=fila_v, column=10).number_format = "#,##0.00 €"
            fila_v += 1
        elif "Compras" in tipo:
            ws_gastos.append(fila_datos)
            ws_gastos.cell(row=fila_g, column=5).number_format = "#,##0.00 €"
            ws_gastos.cell(row=fila_g, column=7).number_format = "#,##0.00 €"
            ws_gastos.cell(row=fila_g, column=9).number_format = "#,##0.00 €"
            ws_gastos.cell(row=fila_g, column=10).number_format = "#,##0.00 €"
            fila_g += 1

    for s in [ws_ventas, ws_gastos]:
        for col in s.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            s.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 4, 13)

    ruta = os.path.join(get_ruta_plantillas(), f"Gestoria_{anio}_{trimestre}T.xlsx")
    wb.save(ruta)
    abrir_archivo_si_escritorio(os.path.abspath(ruta))
    return ruta


def exportar_informe_gestoria_html(anio: int, trimestre: int) -> str:
    """Exporta el informe contable completo en HTML para impresión o envío."""
    movimientos, totales = obtener_resumen_trimestre_db(anio, trimestre)

    filas_ventas_html = ""
    filas_gastos_html = ""

    for m in movimientos:
        num_doc, tipo, fecha = m[1], m[2], m[3][:10]
        cli = html.escape(str(m[4] or "-"))
        conc = html.escape(str(m[5] or "-"))
        base, tipo_iva, c_iva, c_irpf, total = m[6] or 0.0, m[7] or 21.0, m[8] or 0.0, m[10] or 0.0, m[12] or 0.0

        tr = f"""<tr>
            <td style="font-weight:bold;">{num_doc}</td>
            <td>{fecha}</td>
            <td>{cli}</td>
            <td>{conc}</td>
            <td style="text-align:right;">{base:,.2f} €</td>
            <td style="text-align:center;">{tipo_iva:.0f}%</td>
            <td style="text-align:right;">+{c_iva:,.2f} €</td>
            <td style="text-align:right;">-{c_irpf:,.2f} €</td>
            <td style="text-align:right; font-weight:bold;">{total:,.2f} €</td>
        </tr>"""

        if "Facturas" in tipo or "Ventas" in tipo:
            filas_ventas_html += tr
        elif "Compras" in tipo:
            filas_gastos_html += tr

    plantilla = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<title>Cierre Gestoría {trimestre}T - {anio}</title>
<style>
    body {{ font-family: Arial, sans-serif; margin: 30px; color: #0F172A; background-color: #FAFAFA; font-size: 12px; }}
    .header {{ border-bottom: 3px solid #0284C7; padding-bottom: 12px; margin-bottom: 20px; }}
    .header h1 {{ margin: 0; font-size: 20px; color: #0F172A; text-transform: uppercase; }}
    .header .sub {{ font-size: 13px; color: #64748B; margin-top: 4px; }}
    .resumen-fiscal {{ display: table; width: 100%; margin: 20px 0; border-spacing: 12px 0; }}
    .caja-fiscal {{ display: table-cell; width: 50%; border: 1px solid #CBD5E1; border-radius: 8px; padding: 15px; background: #FFFFFF; }}
    .caja-fiscal h3 {{ margin: 0 0 10px 0; font-size: 14px; color: #1E293B; }}
    .caja-fiscal .fila {{ display: flex; justify-content: space-between; margin-bottom: 5px; font-size: 12px; }}
    .caja-fiscal .monto {{ font-size: 17px; font-weight: bold; margin-top: 10px; border-top: 1px solid #E2E8F0; padding-top: 8px; text-align: right; }}
    h2 {{ font-size: 14px; margin-top: 30px; color: #0284C7; border-bottom: 1px solid #CBD5E1; padding-bottom: 5px; }}
    table {{ width: 100%; border-collapse: collapse; margin-top: 10px; font-size: 11px; background: #FFFFFF; }}
    th, td {{ border: 1px solid #E2E8F0; padding: 7px; text-align: left; }}
    th {{ background-color: #1E293B; color: #FFFFFF; font-weight: 600; }}
    .footer {{ margin-top: 40px; font-size: 11px; color: #94A3B8; text-align: center; }}
</style>
</head>
<body>
    <div class="header">
        <h1>Informe Trimestral para Gestoría / Asesoría</h1>
        <div class="sub"><strong>Periodo:</strong> {trimestre}º Trimestre {anio} &nbsp;|&nbsp; <strong>Software:</strong> LZ79 Essential &nbsp;|&nbsp; <strong>Fecha:</strong> {datetime.now().strftime('%d/%m/%Y %H:%M')}</div>
    </div>

    <div class="resumen-fiscal">
        <div class="caja-fiscal">
            <h3>📊 Resumen Modelo 303 (IVA)</h3>
            <div class="fila"><span>Base Ventas / Ingresos:</span> <strong>+{totales['base_ventas']:,.2f} €</strong></div>
            <div class="fila"><span>IVA Repercutido (+):</span> <strong style="color:#0284C7;">+{totales['iva_repercutido']:,.2f} €</strong></div>
            <div class="fila"><span>Base Gastos / Compras:</span> <strong>-{totales['base_gastos']:,.2f} €</strong></div>
            <div class="fila"><span>IVA Soportado (-):</span> <strong style="color:#64748B;">-{totales['iva_soportado']:,.2f} €</strong></div>
            <div class="monto" style="color: {'#16A34A' if totales['modelo_303']<=0 else '#DC2626'};">
                Liquidación Estimada: {totales['modelo_303']:,.2f} €
            </div>
        </div>
        <div class="caja-fiscal">
            <h3>🧮 Resumen Modelo 130 (IRPF)</h3>
            <div class="fila"><span>Ingresos Computables:</span> <strong>+{totales['base_ventas']:,.2f} €</strong></div>
            <div class="fila"><span>Gastos Deducibles:</span> <strong>-{totales['base_gastos']:,.2f} €</strong></div>
            <div class="fila"><span>Rendimiento Neto:</span> <strong>{totales['rendimiento_neto']:,.2f} €</strong></div>
            <div class="fila"><span>Retenciones en Facturas (-):</span> <strong>-{totales['retenciones_ventas']:,.2f} €</strong></div>
            <div class="monto" style="color: #0284C7;">
                Pago Fraccionado (20%): {totales['modelo_130']:,.2f} €
            </div>
        </div>
    </div>

    <h2>1. Libro Registro de Facturas Emitidas (Ventas / Ingresos)</h2>
    <table>
        <thead>
            <tr><th>Nº Factura</th><th>Fecha</th><th>Cliente</th><th>Concepto</th><th style="text-align:right;">Base</th><th style="text-align:center;">IVA</th><th style="text-align:right;">Cuota IVA</th><th style="text-align:right;">IRPF</th><th style="text-align:right;">Total</th></tr>
        </thead>
        <tbody>
            {filas_ventas_html if filas_ventas_html else "<tr><td colspan='9' style='text-align:center; color:#94A3B8;'>No hay facturas emitidas en este trimestre.</td></tr>"}
        </tbody>
    </table>

    <h2>2. Libro Registro de Facturas Recibidas (Compras / Gastos Deducibles)</h2>
    <table>
        <thead>
            <tr><th>Nº Registro</th><th>Fecha</th><th>Proveedor</th><th>Concepto</th><th style="text-align:right;">Base</th><th style="text-align:center;">IVA</th><th style="text-align:right;">Cuota IVA</th><th style="text-align:right;">IRPF</th><th style="text-align:right;">Total</th></tr>
        </thead>
        <tbody>
            {filas_gastos_html if filas_gastos_html else "<tr><td colspan='9' style='text-align:center; color:#94A3B8;'>No hay gastos registrados en este trimestre.</td></tr>"}
        </tbody>
    </table>

    <div class="footer">Documento contable generado automáticamente por LZ79 Essential.</div>
</body>
</html>"""

    ruta = os.path.join(get_ruta_plantillas(), f"Gestoria_{anio}_{trimestre}T.html")
    with open(ruta, "w", encoding="utf-8") as f:
        f.write(plantilla)

    abrir_archivo_si_escritorio(os.path.abspath(ruta))
    return ruta
